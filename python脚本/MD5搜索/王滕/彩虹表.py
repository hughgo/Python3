import hashlib
from openpyxl import Workbook

def md5(i):
    hl = hashlib.md5()
# 重复调用update(arg)方法，是会将传入的arg参数进行拼接，而不是覆盖,m.update(a); m.update(b) 等价于m.update(a+b)
    hl.update(i.encode(encoding='utf-8'))
    return hl.hexdigest()

def sha1(n):
    sha1 = hashlib.sha1()
    sha1.update(n.encode(encoding='utf-8'))
    return sha1.hexdigest()

def save_xlsx():
    filename = '彩虹表.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Top1000彩虹表'
    ws['A1'] = '口令'
    ws['B1'] = 'MD5加密(32位)'
    ws['C1'] = 'sha1加密'
    handle = open('TOP1000.txt', 'r')
    content = handle.readlines()
    for i in content:
        num = content.index(i)
        i = i.strip('\n')
        col_A = 'A%s' % (num + 2)
        col_B = 'B%s' % (num + 2)
        col_C = 'C%s' % (num + 2)
        ws[col_A] = i
        ws[col_B] = md5(i)
        ws[col_C] = sha1(i)

    handle.close()
    wb.save(filename=filename)
    print('保存成功！')

if __name__ == '__main__':
    save_xlsx()


# filename = '彩虹表1.xlsx'
# wb = Workbook()
# ws = wb.active
# ws.title = 'Top1000彩虹表'
# ws['A1'] = '口令'
# ws['B1'] = 'MD5加密(32位)'
# ws['C1'] = 'sha1加密'
# hl = hashlib.md5()
# sha1 = hashlib.sha1()
# handle = open('TOP1000.txt', 'r')
# content = handle.readlines()
# for i in content:
#     num = content.index(i)
#     i = i.strip('\n')
#     hl.update(i.encode(encoding='utf-8'))
#     sha1.update(i.encode(encoding='utf-8'))
#     col_A = 'A%s' % (num + 2)
#     col_B = 'B%s' % (num + 2)
#     col_C = 'C%s' % (num + 2)
#     ws[col_A] = i
#     ws[col_B] = hl.hexdigest()
#     ws[col_C] = sha1.hexdigest()
#
# handle.close()
# wb.save(filename=filename)
# print('保存成功！')